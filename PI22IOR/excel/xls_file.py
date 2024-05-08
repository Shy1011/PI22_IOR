import openpyxl
import os

class Xls_File:
    def __init__(self, file_path):
        self.file_path = file_path

    def xls_open(self):
        # global wb
        if os.path.exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
        else:
            self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        # return goal_sheet

    def xls_write(self, i, j, xxxx):
        self.ws.cell(row=i, column=j, value=xxxx)  # goalworksheet['A1'] = 'Hello, world!'

    def xls_value(self, i, j):              # 获取特定单元格的值
        cell_value = self.ws.cell(row=i, column=j).value
        return cell_value

    def xls_append_row(self):
        row_site = self.ws.max_row + 1
        return row_site

    def xls_append_column(self):
        column_site = self.ws.max_column + 1
        return column_site

    def xls_close(self):
        self.wb.save(self.file_path)
        self.wb.close()
        print(os.path.abspath(self.file_path))
        
    def xls_save(self):
        self.wb.save(self.file_path)

    def xls_create_sheet(self, pSheetName):
        if pSheetName in self.wb:
            self.sheet = self.wb[pSheetName]
        else:
            self.sheet = self.wb.create_sheet(pSheetName)
        return self.sheet

    def xls_active_sheet(self):
        self.wb.active = self.sheet
        self.ws = self.wb.active


if __name__ == '__main__':
    ins_xls_ = Xls_File("../tset123.xlsx")
    ins_xls_.xls_open()

    ins_xls_.xls_write(1, ins_xls_.xls_append_column(), 'a')


    ins_xls_.xls_close()